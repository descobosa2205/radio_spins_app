"""Importación de PLANOS DE BUTACAS desde Excel (motor puro, sin Flask/BD, como seatmap_calc.py).

Convierte un libro .xlsx donde el plano está "dibujado" en celdas (cada celda con un número =
una butaca con ESE número impreso; celda en blanco = hueco/pasillo; textos = etiquetas) en
BLOQUES listos para convertirse en secciones `grid` del diseñador (venue_map.js):

- Cada hoja puede contener VARIAS gradas apiladas (bandas de filas separadas por ≥3 filas vacías)
  y cada banda puede dividirse en varios bloques por sus CABECERAS (celdas combinadas con texto
  encima de la banda: título + «SECTOR N»). Sin cabeceras, se separa por huecos de columnas.
- Las etiquetas de fila (F16, FILA 3, 12, A…) se buscan como el texto más cercano a izquierda/
  derecha de cada fila. Si descienden hacia abajo (F16 arriba … F1 abajo) el bloque sale con
  `row_dir='desc'` (la fila 1 es la de ABAJO), conservando la forma del dibujo sin espejarla.
- NUNCA se interpolan números: en cada fila se intenta encajar la numeración aritmética del
  diseñador (inicio/consecutivos-pares-impares/sentido, con la política de hueco de la sección);
  las filas que no encajan salen con el número EXACTO de cada butaca en `num_overrides`.
- Celdas combinadas grandes CON texto y SIN números dentro (p. ej. «PALCO VIP») se devuelven como
  `labels` (zonas etiquetadas sin butacas); las combinadas vacías (decoración) se ignoran.

API: parse_seatmap_workbook(data: bytes) -> {"sheets": [...], "warnings": [...], "total_seats": N}
Cada bloque replica los campos de una sección grid (rows/cols, gaps por fila, row_seps, num/
row_nums/num_overrides/gap_policy, row_start/row_scheme/row_dir) más su rectángulo Excel
(row_span/col_span) para que el JS coloque los bloques conservando la composición de la hoja.
"""

import re
import unicodedata
from io import BytesIO

# Separaciones que parten el plano: ≥3 filas vacías = otra grada (los pasillos horizontales
# dibujados son de 1-2 filas); ≥3 columnas vacías = otro bloque (los pasillos verticales
# dibujados son de 1-2 columnas).
BAND_ROW_GAP = 3
CLUSTER_COL_GAP = 3
# Ventana (en columnas) donde buscar la etiqueta de fila junto al borde del bloque.
LABEL_WINDOW = 8

_LABEL_RE = re.compile(r"^([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]{0,4})\.?\s*0*(\d+)$")
_ALPHA_RE = re.compile(r"^([A-Za-z]{1,2})$")


def _fold(text):
    s = unicodedata.normalize("NFD", str(text or ""))
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn").lower()


def _seat_value(v):
    """Valor de celda → número de butaca (int ≥ 0) o None si no es una butaca."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            iv = int(v)
        except (ValueError, OverflowError):
            return None
        return iv if float(iv) == float(v) and iv >= 0 else None
    txt = str(v).strip()
    if txt.isdigit():
        return int(txt)
    return None


def _row_label_value(text):
    """Texto de etiqueta de fila → (prefijo, número, esquema) o None.
    'F16' → ('F', 16, 'num') · '12' → ('', 12, 'num') · 'C' → ('', 3, 'alpha')."""
    t = str(text or "").strip()
    m = _LABEL_RE.match(t)
    if m:
        return (m.group(1) or "", int(m.group(2)), "num")
    m = _ALPHA_RE.match(t)
    if m:
        n = 0
        for ch in m.group(1).upper():
            n = n * 26 + (ord(ch) - 64)
        return ("", n, "alpha")
    return None


def _fit_row_numbering(numbers_by_slot, cols, policy):
    """Intenta reproducir los números de UNA fila con la numeración aritmética del diseñador.

    numbers_by_slot: {slot 1-based: número impreso} (solo butacas existentes; el resto son huecos).
    policy: 'skip' (el hueco consume número) o 'renumber' (no consume).
    Devuelve {'start','mode','dir'} si TODOS los números salen exactos, o None."""
    if not numbers_by_slot:
        return None
    vals = list(numbers_by_slot.values())
    if all(v % 2 == 0 for v in vals):
        mode = "even"
    elif all(v % 2 == 1 for v in vals):
        mode = "odd"
    else:
        mode = "seq"
    step = 2 if mode != "seq" else 1
    for dirn in ("ltr", "rtl"):
        order = range(1, cols + 1) if dirn == "ltr" else range(cols, 0, -1)
        k = 0            # butacas (y huecos que consumen) ya contados
        start = None
        ok = True
        for slot in order:
            if slot in numbers_by_slot:
                expected = numbers_by_slot[slot]
                if start is None:
                    start = expected - k * step
                elif start + k * step != expected:
                    ok = False
                    break
                k += 1
            elif policy == "skip":
                k += 1
        # start < 0 no es representable en la UI; en modos par/impar la paridad ya cuadra
        # (start = número_par − 2k). Con un solo número, ltr basta (rtl daría lo mismo).
        if ok and start is not None and start >= 0:
            return {"start": start, "mode": mode, "dir": dirn}
    return None


def _detect_row_labels(block_rows, col_min, col_max, texts_by_row, warnings, block_name):
    """Etiqueta de cada fila del bloque: el texto con pinta de etiqueta más cercano al borde
    izquierdo (validado contra el derecho si existe). Devuelve {excel_row: (prefijo, n, esquema)}."""
    out = {}
    for r in block_rows:
        row_texts = texts_by_row.get(r) or []
        left = right = None
        for c, t in row_texts:
            v = _row_label_value(t)
            if not v:
                continue
            if col_min - LABEL_WINDOW <= c < col_min and (left is None or c > left[0]):
                left = (c, v)
            elif col_max < c <= col_max + LABEL_WINDOW and (right is None or c < right[0]):
                right = (c, v)
        if left and right and left[1] != right[1]:
            warnings.append("«%s»: la etiqueta de fila difiere entre izquierda (%s) y derecha (%s); se usa la izquierda."
                            % (block_name, left[1][0] + str(left[1][1]), right[1][0] + str(right[1][1])))
        v = (left or right or (None, None))[1]
        if v:
            out[r] = v
    return out


def _infer_row_scheme(labels_by_idx, n_rows):
    """De las etiquetas por fila interna (1=arriba) deduce (row_start, row_scheme, row_dir, prefijo)
    o None si no son consecutivas. row_dir 'desc' = la fila 1 (row_start) es la de ABAJO."""
    known = {i: labels_by_idx[i] for i in labels_by_idx}
    if not known:
        return None
    schemes = {v[2] for v in known.values()}
    if len(schemes) != 1:
        return None
    scheme = schemes.pop()
    prefixes = [v[0] for v in known.values() if v[0]]
    prefix = prefixes[0] if prefixes and all(p.upper() == prefixes[0].upper() for p in prefixes) else ""
    # Ascendente hacia abajo: etiqueta(i) = row_start − 1 + i  →  n − i constante.
    asc = {v[1] - i for i, v in known.items()}
    if len(asc) == 1:
        start = asc.pop() + 1
        if start >= 0:
            return (start, scheme, "", prefix)
    # Descendente hacia abajo: etiqueta(i) = row_start + rows − i  →  n + i constante.
    desc = {v[1] + i for i, v in known.items()}
    if len(desc) == 1:
        start = desc.pop() - n_rows
        if start >= 0:
            return (start, scheme, "desc", prefix)
    return None


def _build_block(name, alias, block_seats, warnings, texts_by_row):
    """Convierte las celdas-butaca {(fila, col): número} de un bloque en el dict del bloque."""
    rows_e = sorted({r for (r, _c) in block_seats})
    col_min = min(c for (_r, c) in block_seats)
    col_max = max(c for (_r, c) in block_seats)
    n_rows = len(rows_e)
    cols = col_max - col_min + 1
    row_idx_of = {r: i + 1 for i, r in enumerate(rows_e)}   # fila interna 1 = la de ARRIBA

    # Pasillos horizontales: cada fila Excel vacía entre dos filas del bloque = un rowSep
    # (a todo lo ancho) tras la fila interna de arriba.
    row_seps = []
    for i in range(n_rows - 1):
        for _blank in range(rows_e[i + 1] - rows_e[i] - 1):
            row_seps.append(i + 1)

    # Huecos y números por fila interna.
    numbers = {}   # fila interna → {slot: número}
    for (r, c), n in block_seats.items():
        numbers.setdefault(row_idx_of[r], {})[c - col_min + 1] = n
    gaps = {}
    for ri in range(1, n_rows + 1):
        missing = [s for s in range(1, cols + 1) if s not in (numbers.get(ri) or {})]
        if missing:
            gaps[str(ri)] = missing

    # Etiquetas de fila (F16…F1) → row_start/row_scheme/row_dir.
    labels_e = _detect_row_labels(rows_e, col_min, col_max, texts_by_row, warnings, name)
    labels_by_idx = {row_idx_of[r]: v for r, v in labels_e.items()}
    inferred = _infer_row_scheme(labels_by_idx, n_rows)
    if inferred:
        row_start, row_scheme, row_dir, row_prefix = inferred
    else:
        row_start, row_scheme, row_dir, row_prefix = 1, "num", "", ""
        if labels_by_idx:
            warnings.append("«%s»: las etiquetas de fila no son consecutivas; las filas se numeran 1..%d."
                            % (name, n_rows))

    # Numeración: política de hueco que encaja en MÁS filas; el resto, número exacto por butaca.
    fits = {}
    for policy in ("skip", "renumber"):
        fits[policy] = {ri: _fit_row_numbering(numbers[ri], cols, policy) for ri in numbers}
    n_ok = {p: sum(1 for v in fits[p].values() if v) for p in fits}
    policy = "skip" if n_ok["skip"] >= n_ok["renumber"] else "renumber"
    fitted = fits[policy]

    # Ajuste general de la sección = el patrón más repetido; las demás filas encajadas, por fila.
    counts = {}
    for cfg in fitted.values():
        if cfg:
            key = (cfg["start"], cfg["mode"], cfg["dir"])
            counts[key] = counts.get(key, 0) + 1
    if counts:
        base_key = max(counts, key=lambda k: counts[k])
        num_cfg = {"start": base_key[0], "mode": base_key[1], "dir": base_key[2]}
    else:
        num_cfg = {"start": 1, "mode": "seq", "dir": "ltr"}
    row_nums = {}
    overrides = {}
    for ri in sorted(numbers):
        cfg = fitted.get(ri)
        if cfg:
            if (cfg["start"], cfg["mode"], cfg["dir"]) != (num_cfg["start"], num_cfg["mode"], num_cfg["dir"]):
                row_nums[str(ri)] = cfg
        else:
            for slot, n in numbers[ri].items():
                overrides["%d|%d" % (ri, slot)] = str(n)

    return {
        "name": name,
        "alias": alias,
        "rows": n_rows,
        "cols": cols,
        "row_span": [rows_e[0], rows_e[-1]],
        "col_span": [col_min, col_max],
        "row_start": row_start,
        "row_scheme": row_scheme,
        "row_dir": row_dir,
        "row_prefix": row_prefix,
        "gap_policy": policy,
        "num": num_cfg,
        "row_nums": row_nums,
        "num_overrides": overrides,
        "gaps": gaps,
        "row_seps": row_seps,
        "seat_count": len(block_seats),
        "rows_fitted": sum(1 for v in fitted.values() if v),
        "rows_overridden": sum(1 for v in fitted.values() if not v),
    }


def _column_clusters(seat_cols):
    """Agrupa columnas con butacas en tramos: un hueco de ≥CLUSTER_COL_GAP columnas separa bloques."""
    cols = sorted(seat_cols)
    clusters = []
    for c in cols:
        if clusters and c - clusters[-1][-1] <= CLUSTER_COL_GAP:
            clusters[-1].append(c)
        else:
            clusters.append([c])
    return [(cl[0], cl[-1]) for cl in clusters]


def _parse_sheet(ws, warnings):
    """Una hoja → {"name", "blocks": [...], "labels": [...]} o None si no tiene butacas."""
    seats = {}          # (fila, col) → número
    texts = {}          # (fila, col) → texto (celdas no numéricas)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            n = _seat_value(v)
            if n is not None:
                seats[(cell.row, cell.column)] = n
            else:
                t = str(v).strip()
                if t:
                    texts[(cell.row, cell.column)] = t
    if not seats:
        return None

    texts_by_row = {}
    for (r, c), t in texts.items():
        texts_by_row.setdefault(r, []).append((c, t))

    # Celdas combinadas CON texto (las vacías son decoración y se ignoran).
    text_merges = []
    for m in ws.merged_cells.ranges:
        t = texts.get((m.min_row, m.min_col))
        if t:
            text_merges.append({"r1": m.min_row, "r2": m.max_row, "c1": m.min_col, "c2": m.max_col, "text": t})

    # Bandas de filas: filas con butacas separadas por < BAND_ROW_GAP filas vacías.
    data_rows = sorted({r for (r, _c) in seats})
    bands = []
    for r in data_rows:
        if bands and r - bands[-1][-1] < BAND_ROW_GAP + 1:
            bands[-1].append(r)
        else:
            bands.append([r])

    blocks = []
    header_merge_ids = set()
    prev_end = 0
    for band in bands:
        band_rows = set(range(band[0], band[-1] + 1))
        band_seats = {(r, c): n for (r, c), n in seats.items() if r in band_rows}
        # Cabeceras del bloque: merges de texto entre la banda anterior y esta, que no sean
        # etiquetas de fila. Se agrupan por solape de columnas (título + SECTOR juntos).
        zone = [m for m in text_merges
                if prev_end < m["r1"] < band[0] and not _row_label_value(m["text"])]
        zone.sort(key=lambda m: (m["c1"], m["r1"]))
        groups = []
        for m in zone:
            placed = False
            for g in groups:
                if m["c1"] <= g["c2"] and m["c2"] >= g["c1"]:
                    g["c1"] = min(g["c1"], m["c1"])
                    g["c2"] = max(g["c2"], m["c2"])
                    g["texts"].append((m["r1"], m["text"]))
                    placed = True
                    break
            if not placed:
                groups.append({"c1": m["c1"], "c2": m["c2"], "texts": [(m["r1"], m["text"])]})
        for m in zone:
            header_merge_ids.add(id(m))

        claimed = set()
        band_blocks = []
        for g in sorted(groups, key=lambda g: g["c1"]):
            in_range = {(r, c): n for (r, c), n in band_seats.items() if g["c1"] <= c <= g["c2"]}
            if not in_range:
                continue
            lines = [t for _r, t in sorted(g["texts"])]
            name = lines[0]
            alias = ", ".join(lines[1:])
            claimed.update(in_range)
            band_blocks.append((name, alias, in_range))
        rest = {k: v for k, v in band_seats.items() if k not in claimed}
        if rest:
            for c1, c2 in _column_clusters({c for (_r, c) in rest}):
                chunk = {(r, c): n for (r, c), n in rest.items() if c1 <= c <= c2}
                band_blocks.append((ws.title, "", chunk))
        for name, alias, chunk in band_blocks:
            if len(chunk) < 2:
                warnings.append("Hoja «%s»: se ignora una celda numérica suelta (fila %d)."
                                % (ws.title, min(r for (r, _c) in chunk)))
                continue
            blocks.append(_build_block(name, alias, chunk, warnings, texts_by_row))
        prev_end = band[-1]

    # Nombres únicos dentro de la hoja (varios bloques sin cabecera comparten el de la hoja).
    seen = {}
    for b in blocks:
        seen.setdefault(b["name"], []).append(b)
    for name, same in seen.items():
        if len(same) > 1:
            for i, b in enumerate(same, 1):
                b["name"] = "%s %d" % (name, i)

    # Zonas etiquetadas sin butacas (p. ej. «PALCO VIP»): merges grandes con texto, fuera de las
    # cabeceras, sin ninguna butaca dentro de su rectángulo.
    labels = []
    for m in text_merges:
        if id(m) in header_merge_ids or _row_label_value(m["text"]):
            continue
        w = m["c2"] - m["c1"] + 1
        h = m["r2"] - m["r1"] + 1
        if w * h < 6:
            continue
        if any(m["r1"] <= r <= m["r2"] and m["c1"] <= c <= m["c2"] for (r, c) in seats):
            continue
        labels.append({"text": m["text"], "row_span": [m["r1"], m["r2"]], "col_span": [m["c1"], m["c2"]]})

    return {"name": ws.title, "blocks": blocks, "labels": labels}


def parse_seatmap_workbook(data: bytes) -> dict:
    """Punto de entrada: bytes del .xlsx → hojas con bloques/etiquetas + avisos + total."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Falta la dependencia openpyxl en el servidor.")
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception:
        raise ValueError("No se pudo leer el archivo. ¿Es un Excel (.xlsx) válido?")
    warnings = []
    sheets = []
    for ws in wb.worksheets:
        # La primera hoja de un export de Numbers es un índice («Resumen de exportación»).
        if _fold(ws.title).strip() == "resumen de exportacion":
            continue
        try:
            parsed = _parse_sheet(ws, warnings)
        except Exception:
            warnings.append("Hoja «%s»: no se pudo interpretar; se omite." % ws.title)
            continue
        if parsed and parsed["blocks"]:
            sheets.append(parsed)
    if not sheets:
        raise ValueError("No se encontró ningún plano: ninguna hoja tiene celdas con números de butaca.")
    total = sum(b["seat_count"] for sh in sheets for b in sh["blocks"])
    return {"sheets": sheets, "warnings": warnings, "total_seats": total}
