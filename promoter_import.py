"""Importar TERCEROS desde un fichero (Excel o CSV): motor puro.

Aquí no hay Flask ni base de datos: solo leer el fichero, **detectar sus columnas** y normalizar
los valores. Quién existe ya, qué se crea y qué se actualiza lo decide `app.py`, que es quien puede
mirar la base de datos.

Reglas de la casa que se aplican aquí:
- La cabecera **no tiene por qué estar en la primera fila** (los ficheros de gestoría suelen traer
  un título y una fila en blanco delante): se busca la primera fila que tenga al menos dos celdas
  con texto y que reconozca algún campo, y si no se reconoce ninguno se coge la primera fila con
  varias celdas con texto.
- Los rótulos se comparan **sin acentos, sin puntuación y sin mayúsculas** (`norm_header`), así
  «N.º de C.I.F.» y «nif» son lo mismo sin tener que enumerarlos.
- Lo que NO se reconoce **no se calla ni se tira**: se devuelve con `field=None` para que la
  pantalla pregunte a qué campo de la ficha corresponde.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata

# ── Campos de la ficha del tercero a los que se puede volcar una columna ─────────────────────────
# (clave, etiqueta, tipo, alias). El tipo decide cómo se normaliza y cómo se compara.
FIELDS: list[tuple[str, str, str, tuple[str, ...]]] = [
    ("nick", "Nick (cómo lo llamamos)", "text",
     ("nick", "alias", "nombre comercial", "razon social", "razon", "denominacion", "empresa",
      "proveedor", "cliente", "tercero", "nombre completo", "titular")),
    ("first_name", "Nombre", "text", ("nombre", "nombre de pila", "first name", "name")),
    ("last_name", "Apellidos", "text", ("apellidos", "apellido", "apellido1", "apellidos1",
                                        "last name", "surname")),
    ("tax_id", "DNI / NIF / CIF", "tax_id", ("dni", "nif", "cif", "nie", "documento",
                                             "n documento", "num documento", "numero documento",
                                             "identificacion", "vat", "nif cif", "cif nif",
                                             "dni cif", "id fiscal")),
    ("contact_email", "Email", "email", ("email", "e mail", "correo", "correo electronico", "mail",
                                         "email contacto", "correo contacto")),
    ("contact_phone", "Teléfono", "phone", ("telefono", "tlf", "tel", "movil", "celular", "phone",
                                            "telefono contacto", "telefono movil")),
    ("address", "Domicilio", "text", ("domicilio", "direccion", "direccion postal", "address",
                                      "domicilio particular")),
    ("fiscal_address", "Dirección fiscal (calle)", "text",
     ("direccion fiscal", "domicilio fiscal", "calle", "via", "domicilio social")),
    ("fiscal_postal_code", "Código postal", "postal", ("codigo postal", "cp", "c postal", "zip",
                                                       "postal", "codpostal")),
    ("fiscal_city", "Municipio", "text", ("municipio", "ciudad", "poblacion", "localidad", "city")),
    ("fiscal_province", "Provincia", "text", ("provincia", "province", "estado")),
    ("fiscal_country", "País", "text", ("pais", "country", "nacion")),
    ("bank_account", "IBAN / cuenta bancaria", "iban", ("iban", "cuenta", "cuenta bancaria",
                                                        "numero de cuenta", "n cuenta", "ccc",
                                                        "banco cuenta")),
    ("bank_bic", "SWIFT / BIC", "text", ("bic", "swift", "swift bic", "bic swift", "codigo bic")),
    ("kind", "Tipo (empresa / institución / persona)", "kind",
     ("tipo", "tipo de tercero", "clase", "categoria", "tipo cliente", "tipo proveedor")),
    ("prl_type", "Cómo factura (autónomo / alta puntual / empresa)", "prl_type",
     ("regimen", "tipo de alta", "prl", "autonomo", "situacion laboral", "como factura")),
    ("hotel_notes", "Petición de hoteles", "text", ("hotel", "hoteles", "peticion hotel",
                                                    "preferencias hotel")),
    ("travel_notes", "Necesidades de viaje", "text", ("viaje", "viajes", "necesidades de viaje",
                                                      "preferencias de viaje")),
]

FIELD_LABELS = {key: label for key, label, _kind, _alias in FIELDS}
FIELD_KINDS = {key: kind for key, _label, kind, _alias in FIELDS}
FIELD_KEYS = [key for key, _l, _k, _a in FIELDS]

# Destinos especiales de una columna que no es ninguno de los campos de arriba.
TARGET_IGNORE = "__ignore__"
TARGET_ALT = "__alt__"          # se guarda como dato extra CON EL NOMBRE DE LA COLUMNA

_KIND_VALUES = {
    "empresa": {"empresa", "sociedad", "sl", "s l", "sa", "s a", "company", "juridica",
                "persona juridica", "mercantil"},
    "institucion": {"institucion", "ayuntamiento", "organismo", "publica", "administracion",
                    "diputacion", "consejeria", "ministerio", "universidad"},
}
_PRL_VALUES = {
    "AUTONOMO": {"autonomo", "autonoma", "freelance", "ret", "reta", "autonomos"},
    "PUNTUAL": {"puntual", "alta puntual", "artista", "por dias", "eventual"},
    "EMPRESA": {"empresa", "fijo", "plantilla", "cuenta ajena", "oficina"},
}


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text or "") if not unicodedata.combining(c))


def norm_header(text) -> str:
    """El rótulo de una columna, comparable: sin acentos, sin puntuación y en minúsculas."""
    value = strip_accents(str(text or "")).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


_ALIAS_RE_CACHE: dict[str, re.Pattern] = {}


def _alias_re(alias: str) -> re.Pattern:
    """El alias con **puntuación permitida entre sus letras** y pegado a un límite de palabra.

    ⚠️ Sin esto, «N.º de C.I.F.» no se reconoce: al normalizar queda «n o de c i f» y ningún alias
    encaja (bug real de la primera prueba). Y con la puntuación permitida hace falta el límite por
    la izquierda: si no, el «nie» de «conveniente» se tomaría por un NIE.
    """
    cached = _ALIAS_RE_CACHE.get(alias)
    if cached is None:
        cuerpo = r"[^a-z0-9]*".join(re.escape(c) for c in norm_header(alias).replace(" ", ""))
        cached = re.compile(r"(?<![a-z])" + cuerpo + r"(?![a-z])")
        _ALIAS_RE_CACHE[alias] = cached
    return cached


def guess_field(header) -> str | None:
    """A qué campo de la ficha corresponde una columna. None = no se reconoce (hay que preguntar)."""
    key = norm_header(header)
    if not key:
        return None
    for field, _label, _kind, aliases in FIELDS:
        for alias in aliases:
            if key == norm_header(alias):
                return field
    # Segunda pasada: el rótulo CONTIENE el alias («nº de teléfono móvil» → telefono, «N.º de
    # C.I.F.» → tax_id). Se prueban los más largos primero para que «direccion fiscal» gane a
    # «direccion».
    plano = strip_accents(str(header or "")).lower()
    candidatos = []
    for field, _label, _kind, aliases in FIELDS:
        for alias in aliases:
            a = norm_header(alias).replace(" ", "")
            if len(a) >= 3 and _alias_re(alias).search(plano):
                candidatos.append((len(a), field))
    if candidatos:
        candidatos.sort(reverse=True)
        return candidatos[0][1]
    return None


def clean_tax_id(value) -> str:
    return re.sub(r"[^0-9A-Z]", "", strip_accents(str(value or "")).upper())


def clean_iban(value) -> str:
    return re.sub(r"[^0-9A-Z]", "", strip_accents(str(value or "")).upper())


def _cell_text(value) -> str:
    """El texto de una celda. ⚠️ Los números de Excel llegan como float: un teléfono o un código
    postal salían «638123456.0» y «41001.0», que no valen para nada."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        value = int(value)
    if isinstance(value, str):
        text = value.strip()
        # ⚠️ Un CSV exportado de Excel trae los números con decimales: un teléfono salía
        # «638123456.0» y un código postal «41001.0» (bug real de la primera prueba).
        # ⚠️⚠️ SOLO UNO O DOS DECIMALES: con `\.0+` esta limpieza se comía los MILES del modelo de
        # euros —«40.000» se quedaba en «40», «1.000» en «1»— y con ello el importe de cualquier
        # fichero que se importara (compradores, terceros, liquidaciones): cuarenta mil euros
        # entraban como cuarenta (bug real de dinero). Tres ceros detrás del punto son un grupo de
        # MILES, no la coma decimal que mete Excel.
        if re.fullmatch(r"-?\d+\.0{1,2}", text):
            text = text.split(".", 1)[0]
        return text
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            return str(value)
    return str(value).strip()


def normalize_value(field: str, value) -> str:
    """El valor tal como se va a guardar en ese campo."""
    text = _cell_text(value)
    if not text:
        return ""
    kind = FIELD_KINDS.get(field, "text")
    if kind == "email":
        return text.replace(" ", "").lower()
    if kind == "tax_id":
        return clean_tax_id(text)
    if kind == "iban":
        return clean_iban(text)
    if kind == "postal":
        digits = re.sub(r"\D", "", text)
        return digits.zfill(5) if 0 < len(digits) <= 5 else (digits or text)
    if kind == "phone":
        cleaned = re.sub(r"[^0-9+]", "", text)
        return cleaned or text
    if kind == "kind":
        key = norm_header(text)
        for target, words in _KIND_VALUES.items():
            if key in {norm_header(w) for w in words}:
                return target
        return ""  # persona / tercero genérico: en la ficha es NULL
    if kind == "prl_type":
        key = norm_header(text)
        for target, words in _PRL_VALUES.items():
            if key in {norm_header(w) for w in words}:
                return target
        return ""
    return re.sub(r"\s+", " ", text).strip()


def _rows_from_xlsx(data: bytes) -> list[list]:
    import openpyxl  # dependencia ya usada por el resto de importadores
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]          # SOLO la primera hoja, como el resto de importadores
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _rows_from_csv(data: bytes) -> list[list]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("No se pudo leer el CSV (codificación desconocida).")
    muestra = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";" if muestra.count(";") > muestra.count(",") else ","
    filas = [list(r) for r in csv.reader(io.StringIO(text), dialect)]
    return _resplit_csv(text, filas, getattr(dialect, "delimiter", ","))


def _resplit_csv(text: str, filas: list, usado: str) -> list:
    """⚠️⚠️ `csv.Sniffer` SE QUEDA CON LA COMA EN UN CSV ESPAÑOL, y ahí la coma es el separador
    DECIMAL: «…;Tema Cinco;120;45,10» se partía en «…;Tema Cinco;120;45» y «10», así que el importe
    salía **10** y el título arrastraba media fila (bug real).

    Si con el separador elegido quedan celdas que TODAVÍA traen dos o más de otro candidato, ese otro
    es el separador de verdad y se relee con él."""
    if not filas:
        return filas
    for otro in (";", "\t", "|", ","):
        if otro == usado:
            continue
        sospechosas = sum(1 for fila in filas[:30] for celda in (fila or [])
                          if isinstance(celda, str) and celda.count(otro) >= 2)
        if not sospechosas:
            continue
        try:
            d = csv.excel()
            d.delimiter = otro
            nuevas = [list(r) for r in csv.reader(io.StringIO(text), d)]
        except Exception:
            continue
        # Solo vale si de verdad parte en MÁS columnas (si no, se deja lo que había).
        if nuevas and max((len(x) for x in nuevas[:30]), default=0) > max((len(x) for x in filas[:30]), default=0):
            return nuevas
    return filas


def _header_index(rows: list[list], guesser=None) -> int:
    """Qué fila es la cabecera. Manda la primera que reconozca algún campo; si ninguna lo hace, la
    primera con dos o más celdas con texto (el fichero se podrá mapear a mano)."""
    guesser = guesser or guess_field
    respaldo = None
    for i, row in enumerate(rows[:30]):
        textos = [_cell_text(c) for c in (row or [])]
        con_texto = [t for t in textos if t]
        if len(con_texto) < 2:
            continue
        if respaldo is None:
            respaldo = i
        if any(guesser(t) for t in con_texto):
            return i
    return respaldo if respaldo is not None else 0


def read_rows(data: bytes, filename: str = "") -> list[list]:
    """Las filas CRUDAS de un Excel o un CSV. Es el lector que comparten los importadores de la
    casa (terceros y compradores): un solo sitio que sabe leer un fichero."""
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt")):
        rows = _rows_from_csv(data)
    elif name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(data)
    else:
        # Sin extensión reconocible: se prueba como Excel y, si no cuela, como CSV.
        try:
            rows = _rows_from_xlsx(data)
        except Exception:
            rows = _rows_from_csv(data)
    if not rows:
        raise ValueError("El fichero está vacío.")
    return rows


def header_index(rows: list[list], guesser=None) -> int:
    """Qué fila es la cabecera (ver `_header_index`), con el reconocedor que se le pase."""
    return _header_index(rows, guesser)


def parse_columns(rows: list[list], guesser=None) -> dict:
    """De las filas crudas a {columns, rows, sheet_rows}, reconociendo cada columna con `guesser`.

    Lo comparten los importadores: lo único que cambia entre uno y otro es a QUÉ campos se puede
    volcar una columna, o sea el reconocedor."""
    guesser = guesser or guess_field
    h = header_index(rows, guesser)
    header = [_cell_text(c) for c in (rows[h] or [])]
    cuerpo = []
    for row in rows[h + 1:]:
        textos = [_cell_text(c) for c in (row or [])]
        if any(textos):
            cuerpo.append(textos)

    ancho = max([len(header)] + [len(r) for r in cuerpo] or [0])
    header += [""] * (ancho - len(header))
    cuerpo = [r + [""] * (ancho - len(r)) for r in cuerpo]

    columns = []
    usados = set()
    for idx in range(ancho):
        rotulo = header[idx]
        muestras = [r[idx] for r in cuerpo[:8] if r[idx]][:3]
        if not rotulo and not muestras:
            continue  # columna completamente vacía: no se enseña
        campo = guesser(rotulo)
        # Un campo no se puede rellenar desde dos columnas a la vez: la segunda se pregunta.
        if campo and campo in usados:
            campo = None
        if campo:
            usados.add(campo)
        columns.append({
            "index": idx,
            "header": rotulo or f"Columna {idx + 1}",
            "field": campo,
            "auto": bool(campo),
            "samples": muestras,
        })
    return {"columns": columns, "rows": cuerpo, "sheet_rows": len(cuerpo)}


def parse_file(data: bytes, filename: str = "") -> dict:
    """Lee el fichero y devuelve sus columnas (con el campo que se les ha reconocido) y sus filas.

    {"columns": [{"index", "header", "field", "auto", "samples": [...]}],
     "rows": [[texto, …], …],          # ya sin la cabecera y sin filas vacías
     "sheet_rows": nº de filas leídas}
    """
    return parse_columns(read_rows(data, filename), guess_field)


def apply_mapping(rows: list[list], mapping: dict) -> list[dict]:
    """Convierte las filas del fichero en fichas: {"values": {campo: valor},
    "alt": [{"label", "value"}]} (los `alt` son las columnas que se han decidido conservar como
    dato extra, con el nombre de su columna)."""
    out = []
    for row in rows or []:
        values, alt = {}, []
        for key, target in (mapping or {}).items():
            try:
                idx = int(key)
            except Exception:
                continue
            if idx < 0 or idx >= len(row):
                continue
            crudo = _cell_text(row[idx])
            if not crudo:
                continue
            destino = (target or {}).get("field") if isinstance(target, dict) else target
            destino = (destino or "").strip()
            if not destino or destino == TARGET_IGNORE:
                continue
            if destino == TARGET_ALT:
                etiqueta = ((target or {}).get("label") if isinstance(target, dict) else "") or ""
                alt.append({"label": (etiqueta or f"Columna {idx + 1}").strip(), "value": crudo})
                continue
            if destino not in FIELD_LABELS:
                continue
            limpio = normalize_value(destino, crudo)
            if limpio:
                values[destino] = limpio
        if values or alt:
            out.append({"values": values, "alt": alt})
    return out
